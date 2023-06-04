
import time
import pandas as pd
import tabulate
from jira import JIRA
import os
from openpyxl import load_workbook
import shortuuid
import sys

from datetime import datetime

api_token=os.environ.get('API')

counter=1
jiradf_backup=None
jira = JIRA(server='https://jiratickettrial.atlassian.net/', basic_auth=('pythonlover260@gmail.com', api_token))

def generate_unique_ids(num_ids):
    return [str(shortuuid.uuid()) for _ in range(num_ids)]

def preserve():
    global jiradf_backup
    excel_file = 'Jira.xlsx'
    workbook = load_workbook(excel_file)
    consolidated_sheet = 'ConsolidatedSheet'
    week_sheetname = [sheet for sheet in workbook.sheetnames if sheet != consolidated_sheet and sheet.startswith('W')]
    jiraticketdata=[sheet for sheet in workbook.sheetnames if sheet.startswith('JiraTicketData')]
    other_sheet_df = {}

    for sheet in workbook.sheetnames:
        if sheet not in [consolidated_sheet]+week_sheetname+jiraticketdata:
            sheet_df=pd.read_excel(excel_file,sheet_name=sheet)
            other_sheet_df[sheet]=sheet_df
        if (jiraticketdata) and (jiradf_backup is None):
            jiradf_backup=pd.read_excel(excel_file,sheet_name=jiraticketdata[0])
    return week_sheetname,other_sheet_df,consolidated_sheet,jiradf_backup

def reload_dataframe():
    global df1,df2  # Declare df as a global variable
    week_sheetname,other_sheet_df,consolidated_sheet,jiradf_backup=preserve()
    if len(week_sheetname) == 0:
        print("No weekly file found")
        sys.exit()

    df1 = pd.read_excel("Jira.xlsx", sheet_name='ConsolidatedSheet')
    df2 = pd.read_excel("Jira.xlsx", sheet_name=week_sheetname[0])
    df2=pd.DataFrame(df2)
    return df1,df2,other_sheet_df

def removeNull(df1,df2):
    df1.dropna(subset=['SystemName'],inplace=True)
    df2.dropna(subset=['SystemName'],inplace=True)
    return df1,df2

def concatdf(df_list,other_sheet_df):

    issueid=[]
    master_df=pd.DataFrame()
    for df in df_list:
        num_rows=len(df)
        unique_ids=generate_unique_ids(num_rows)

        if 'IssueID' in df.columns:
            if df['IssueID'].isnull().any():
                df.loc[df['IssueID'].isnull(),'IssueID']=unique_ids

        else:
            df['IssueID'] = unique_ids


            # else:
            #   max_id=df['IssueID'].max()
            #   missing_id=df['IssueID'].isnull().sum()
            #   new_id=range(max_id+1,max_id+missing_id+1)
            #   df.loc[df["IssueID"].isnull(),'IssueID']=new_id
        master_df=pd.concat([master_df,df])
    duplicate_rows = master_df[master_df.duplicated(subset=['SystemName', 'Vulnerabilityfound', 'Environment'])]

    if not duplicate_rows.empty:
        merged_df_issue_list = duplicate_rows['IssueID'].to_list()
    else:
        merged_df_issue_list=[]

    while True:
        try:
            with pd.ExcelWriter('Jira.xlsx') as writer:
                master_df.to_excel(writer,sheet_name='ConsolidatedSheet',index=False)
                for sheet, sheet_df in other_sheet_df.items():
                    sheet_df.to_excel(writer, sheet_name=sheet, index=False)
            break
        except PermissionError:
            print("Please close the file to proceed")
            time.sleep(5)

    return master_df,merged_df_issue_list


def jirafile(masterdf,merged_df_issue_list):

    count=0
    global jiradf
    global sheet_name
    sheet_name='JiraTicketData'
    jiradf=None

    filtered_list=[]

    if len(merged_df_issue_list)!=0:
        master_issue_id = masterdf['IssueID'].values.tolist()

        for issues in master_issue_id:
            if issues not in merged_df_issue_list:
                if 'TicketID' in masterdf.columns:
                    print(issues)
                    if masterdf.loc[masterdf['IssueID'] == issues, 'TicketID'].isnull().all():
                        filtered_list.append(issues)

                else:
                    masterdf['TicketID'] = float('NaN')
                    filtered_list.append(issues)


        if filtered_list:
            filtered_df=masterdf[masterdf["IssueID"].isin(filtered_list)]
            grouped_df=filtered_df.groupby(['Vulnerabilityfound'],as_index=False)
            jiradf=grouped_df.agg(lambda x:','.join(x.astype(str)))
            jiradf.reset_index(drop=True, inplace=True)  # Reset the index to preserve the original unique IDs


    else:
        grouped_df = masterdf.groupby(['Vulnerabilityfound'], as_index=False)
        jiradf = grouped_df.agg(lambda x: ','.join(x.astype(str).unique()))
        jiradf.reset_index(drop=True, inplace=True)  # Reset the index to preserve the original unique IDs

    week_sheetname, other_sheet_df, consolidated_sheet, jiradf_backup = preserve()

    if len(filtered_list)==0:
        jiradf=jiradf_backup.copy()

    try:
        with pd.ExcelWriter('Jira.xlsx') as writer:
            if jiradf is not None:
                jiradf.to_excel(writer,sheet_name=sheet_name,index=False)
                masterdf.to_excel(writer,sheet_name='ConsolidatedSheet',index=False)

            else:
                masterdf.to_excel(writer, sheet_name='ConsolidatedSheet', index=False)

            for sheet, sheet_df in other_sheet_df.items():
                sheet_df.to_excel(writer, sheet_name=sheet, index=False)


    except PermissionError:
        print("Please close the file to proceed")
        time.sleep(5)
    # except ValueError:
    #     print('JiraTicketData already exists in the file,creating new file')
    #     workbook=load_workbook('Jira.xlsx')
    #     for sheet in workbook.sheetnames:
    #         if sheet.startswith('JiraTicketData'):
    #             count=count+1
    #     sheet_name=f"JiraTicketData{count+1}"
    #     print("New Sheet name",sheet_name)

    return jiradf

def jiraticket(jiradf,master_df):
    global sheet_name
    Prefix1='Additional Details are'
    issue_metadata = jira.createmeta(projectKeys='JIRA')
    priority_mapping = {
    'Highest':'Highest',
    'High': 'High',
    'Low':'Low',
    'Lowest':'Lowest'
    }

    for index,row in jiradf.iterrows():
        issue_identifier = row['IssueID']

        #duedate_str=row['DueDate']
        #print(duedate_str)
        #duedate_dt = pd.to_datetime(duedate_str,format='%Y-%m-%d').to_pydatetime()
        #print(duedate_dt)
        #priority_value = row['Severity'].split(',')
        #priority_data= priority_mapping.get(priority_value[0])

        if issue_identifier:
            ip_addresses = row['IPAddress'].split(',')
            system_names = row['SystemName'].split(',')
            platform= row['Platform'].split(',')
            OSname=row['OSName'].split(',')
            Env = row['Environment'].split(',')
            table_data=[['IPAddress','SystemName','Platform','OSName','Environment']]+list(zip(ip_addresses,system_names,platform,OSname,Env))
            table = tabulate.tabulate(table_data, headers='firstrow', tablefmt='pipe')
            #duedate_dt = pd.to_datetime(duedate_str, format='%Y-%m-%d').to_pydatetime()
            #print(duedate_dt)
            issue_data = {
                'project': issue_metadata['projects'][0]['key'],
                'summary': row['Vulnerabilityfound'],
                'description': '{}\n{}'.format(Prefix1,table),
                'issuetype':{'name':'Task'},
                #'duedate':duedate_dt.isoformat(),
                'assignee': row['Assigned'],
                #'priority':{'name':priority_data}
                }

            issue=jira.create_issue(**issue_data)
        ticket_id=issue.key

        jiradf.at[index, 'TicketID'] = ticket_id

        for index, row in jiradf.iterrows():
            issue_id = row['IssueID']
            issue_list = issue_id.split(',')
            ticket_id = row['TicketID']
            master_df.loc[master_df['IssueID'].isin(issue_list), 'TicketID'] = ticket_id

    week_sheetname, other_sheet_df,consolidated_sheet,jiradf_backup=preserve()

    while True:
        try:
            with pd.ExcelWriter('Jira.xlsx') as writer:
                jiradf.to_excel(writer, sheet_name=sheet_name, index=False)
                masterdf.to_excel(writer,sheet_name=consolidated_sheet,index=False)
                for sheet, sheet_df in other_sheet_df.items():
                    sheet_df.to_excel(writer, sheet_name=sheet, index=False)
            break
        except PermissionError:
            print("Please close the file to proceed")
            time.sleep(5)


df1,df2,other_sheet_df = reload_dataframe()
df1,df2= removeNull(df1,df2)
df_list=[df1,df2]
masterdf,merged_df_issue_list=concatdf(df_list,other_sheet_df)
sheet_name=None
jiradf=jirafile(masterdf,merged_df_issue_list)
if jiradf is not None:
    jiraticket(jiradf,masterdf)
else:
    print("All the weekly issues already have tickets, so no new tickets required")




