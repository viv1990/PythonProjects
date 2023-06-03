
import time
import zipfile

import pandas as pd
import tabulate
from jira import JIRA
import os
from openpyxl import load_workbook
import shortuuid
import sys
import zipfile

from datetime import datetime

api_token=os.environ.get('API')

counter=1
jiradf_backup=None
jira = JIRA(server='https://jiratickettrial.atlassian.net/', basic_auth=('er.viveksoni90@gmail.com', api_token))

def generate_unique_ids(num_ids):
    return [str(shortuuid.uuid()) for _ in range(num_ids)]

def preserve(file,jiradf_backup):
    excel_file = file
    workbook = load_workbook(excel_file)
    consolidated_sheet = 'ConsolidatedSheet'
    week_sheetname = [sheet for sheet in workbook.sheetnames if sheet != consolidated_sheet and sheet.startswith('Week')]
    jiraticketdata=[sheet for sheet in workbook.sheetnames if sheet.startswith('JiraTicketData')]
    other_sheet_df = {}

    for sheet in workbook.sheetnames:
        if sheet not in [consolidated_sheet]+week_sheetname+jiraticketdata:
            sheet_df=pd.read_excel(excel_file,sheet_name=sheet)
            other_sheet_df[sheet]=sheet_df
        if (jiraticketdata) and (jiradf_backup is None):
            jiradf_backup=pd.read_excel(excel_file,sheet_name=jiraticketdata[0])
    return other_sheet_df,jiradf_backup
def reload_dataframe(file):
    global df1,df2  # Declare df as a global variable
    messages=[]
    consolidated_sheet = 'ConsolidatedSheet'
    try:
        workbook = load_workbook(filename=file)
    except zipfile.BadZipfile:
        messages.append("Invalid/Corrupt file. Please upload a valid Excel file.")
        return messages, None, None


    if consolidated_sheet not in workbook.sheetnames:
        messages.append("Consolidated sheet not found, upload a proper file")
        return messages, None, None

    try:
        df1=pd.read_excel(file,sheet_name=consolidated_sheet)
    except Exception as e:
        messages.append("Error loading ConsolidatedSheet: " + str(e))
        return messages, None, None

    week_sheetname = [sheet for sheet in workbook.sheetnames if sheet != consolidated_sheet and sheet.startswith('Week')]
    if len(week_sheetname) == 0:
        if ('IssueID' in df1.columns and 'TicketID' in df1.columns):
            messages.append("No weekly file found, the consolidated tracker already has tickets")
            print(messages)
            return messages,df1,None
        else:
            return messages,df1,None
    else:
        df2 = pd.read_excel(file,sheet_name=week_sheetname[0])
        return messages,df1,df2

def removeNull(df1,df2):

    if df1 is not None and df2 is None:
        df1.dropna(subset=['SystemName'],inplace=True)
        return df1,None
    elif df2 is not None and df1 is None:
        df2.dropna(subset=['SystemName'],inplace=True)
        return None, df2
    elif df1 is None and df2 is None:
        return None, None
    elif df1 is not None and df2 is not None:
        df1.dropna(subset=['SystemName'], inplace=True)
        df2.dropna(subset=['SystemName'], inplace=True)

        return df1, df2

def concatdf(df_list,other_sheet_df):

    issueid=[]
    master_df=pd.DataFrame()
    for df in df_list:
        num_rows=len(df)
        unique_ids=generate_unique_ids(num_rows)

        if 'IssueID' in df.columns:
            if df['IssueID'].isnull().any():
                df.loc[df['IssueID'].isnull(),'IssueID']=unique_ids

        else:
            df['IssueID'] = unique_ids


            # else:
            #   max_id=df['IssueID'].max()
            #   missing_id=df['IssueID'].isnull().sum()
            #   new_id=range(max_id+1,max_id+missing_id+1)
            #   df.loc[df["IssueID"].isnull(),'IssueID']=new_id
        master_df=pd.concat([master_df,df])
    duplicate_rows = master_df[master_df.duplicated(subset=['SystemName', 'Vulnerabilityfound', 'Environment'])]

    if not duplicate_rows.empty:
        merged_df_issue_list = duplicate_rows['IssueID'].to_list()
    else:
        merged_df_issue_list=[]

    while True:
        try:
            with pd.ExcelWriter('Jira.xlsx') as writer:
                master_df.to_excel(writer,sheet_name='ConsolidatedSheet',index=False)
                for sheet, sheet_df in other_sheet_df.items():
                    sheet_df.to_excel(writer, sheet_name=sheet, index=False)
            break
        except PermissionError:
            print("Please close the file to proceed")
            time.sleep(5)

    return master_df,merged_df_issue_list


def jirafile(file_path,master_df,merged_df_issue_list,jiradf_backup,other_sheet_df):
    global jiradf
    global sheet_name
    sheet_name='JiraTicketData'
    jiradf=None

    filtered_list=[]

    if len(merged_df_issue_list)!=0:
        master_issue_id = master_df['IssueID'].values.tolist()
        for issues in master_issue_id:
            if issues not in merged_df_issue_list:
                if 'TicketID' in master_df.columns:
                    if master_df.loc[master_df['IssueID'] == issues, 'TicketID'].isnull().all():
                        print("Issues for filtered list",issues)
                        filtered_list.append(issues)
                        print("My filtered list",filtered_list)

                else:
                    master_df['TicketID'] = float('NaN')
                    print("TicketID not in ConsolidatedSheet",master_df.head(50))
                    filtered_list.append(issues)
                    print("My filtered list if no 'Ticket_Id in ConsolidatedSheet", filtered_list)

        if filtered_list:
            filtered_df=master_df[master_df["IssueID"].isin(filtered_list)]
            grouped_df=filtered_df.groupby(['Vulnerabilityfound'],as_index=False)
            jiradf=grouped_df.agg(lambda x:','.join(x.astype(str)))
            jiradf.reset_index(drop=True, inplace=True)  # Reset the index to preserve the original unique IDs

    else:
        grouped_df = master_df.groupby(['Vulnerabilityfound'], as_index=False)
        jiradf = grouped_df.agg(lambda x: ','.join(x.astype(str)))
        jiradf.reset_index(drop=True, inplace=True)  # Reset the index to preserve the original unique IDs

    if len(filtered_list)==0:
        print("Length of filtered list is 0")
        print(jiradf_backup.head(50))
        jiradf=jiradf_backup.copy()
        print(jiradf.head(50))

    try:
        with pd.ExcelWriter(file_path) as writer:
            if jiradf is not None:
                print("Jiradf is not none")
                print("Inside Writer",jiradf.head(50))
                jiradf.to_excel(writer,sheet_name=sheet_name,index=False)
                master_df.to_excel(writer,sheet_name='ConsolidatedSheet',index=False)

            else:
                master_df.to_excel(writer, sheet_name='ConsolidatedSheet', index=False)

            for sheet, sheet_df in other_sheet_df.items():
                sheet_df.to_excel(writer, sheet_name=sheet, index=False)


    except PermissionError:
        print("Please close the file to proceed")
        time.sleep(5)
    # except ValueError:
    #     print('JiraTicketData already exists in the file,creating new file')
    #     workbook=load_workbook('Jira.xlsx')
    #     for sheet in workbook.sheetnames:
    #         if sheet.startswith('JiraTicketData'):
    #             count=count+1
    #     sheet_name=f"JiraTicketData{count+1}"
    #     print("New Sheet name",sheet_name)

    return jiradf,sheet_name,filtered_list

def jiraticket(sheet_name,jiradf,master_df,other_sheet_df):
    Prefix1='Additional Details are'
    issue_metadata = jira.createmeta(projectKeys='JIR')
    priority_mapping = {
    'Highest':'Highest',
    'High': 'High',
    'Low':'Low',
    'Lowest':'Lowest'
    }

    for index,row in jiradf.iterrows():
        issue_identifier = row['IssueID']

        #duedate_str=row['DueDate']
        #print(duedate_str)
        #duedate_dt = pd.to_datetime(duedate_str,format='%Y-%m-%d').to_pydatetime()
        #print(duedate_dt)
        #priority_value = row['Severity'].split(',')
        #priority_data= priority_mapping.get(priority_value[0])

        if issue_identifier:
            ip_addresses = row['IPAddress'].split(',')
            system_names = row['SystemName'].split(',')
            platform= row['Platform'].split(',')
            OSname=row['OSName'].split(',')
            Env = row['Environment'].split(',')
            table_data=[['IPAddress','SystemName','Platform','OSName','Environment']]+list(zip(ip_addresses,system_names,platform,OSname,Env))
            table = tabulate.tabulate(table_data, headers='firstrow', tablefmt='pipe')
            #duedate_dt = pd.to_datetime(duedate_str, format='%Y-%m-%d').to_pydatetime()
            #print(duedate_dt)
            issue_data = {
                'project': issue_metadata['projects'][0]['key'],
                'summary': row['Vulnerabilityfound'],
                'description': '{}\n{}'.format(Prefix1,table),
                'issuetype':{'name':'Task'},
                #'duedate':duedate_dt.isoformat(),
                'assignee': row['Assigned'],
                #'priority':{'name':priority_data}
                }

            issue=jira.create_issue(**issue_data)
        ticket_id=issue.key

        jiradf.at[index, 'TicketID'] = ticket_id

        for index, row in jiradf.iterrows():
            issue_id = row['IssueID']
            issue_list = issue_id.split(',')
            ticket_id = row['TicketID']
            master_df.loc[master_df['IssueID'].isin(issue_list), 'TicketID'] = ticket_id

    print(jiradf.head(50))

    while True:
        try:
            with pd.ExcelWriter('Jira.xlsx') as writer:
                jiradf.to_excel(writer, sheet_name=sheet_name, index=False)
                master_df.to_excel(writer,sheet_name='ConsolidatedSheet',index=False)
                for sheet, sheet_df in other_sheet_df.items():
                    sheet_df.to_excel(writer, sheet_name=sheet, index=False)
            break
        except PermissionError:
            print("Please close the file to proceed")
            time.sleep(5)


# message,df1,df2= reload_dataframe('Jira.xlsx')
# df1,df2= removeNull(df1,df2)
# df_list=[df1,df2]
# jiradf_backup=None
# other_sheet_df, jiradf_backup = preserve ('Jira.xlsx',jiradf_backup)
# master_df,merged_df_issue_list=concatdf(df_list,other_sheet_df)
# sheet_name=None
# jiradf,sheet_name,filtered_list=jirafile('Jira.xlsx',master_df, merged_df_issue_list,jiradf_backup,other_sheet_df)
# if jiradf is not None:
#     jiraticket(sheet_name,jiradf, master_df,other_sheet_df,)
# else:
#     print("All the weekly issues already have tickets, so no new tickets required")




